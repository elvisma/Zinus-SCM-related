# -*- coding: utf-8 -*-
"""
Created on Tue May  7 22:12:44 2019

@author: Elvis Ma
"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import os
# folder directory
folderpath="C:/Users/Elvis Ma/Desktop/Weekly Work/space_planning/"

startCol=52 # 10/6 2019 weekly change 


# All category SISRs to modify, save

SB_SISR="SISR_SB_AMZ.xlsm"
SB_path=folderpath+SB_SISR

BX_SISR="SISR_BX_AMZ.xlsm"
BX_path=folderpath+BX_SISR

FR_SISR="SISR_FR_AMZ.xlsm"
FR_path=folderpath+FR_SISR

FM_SISR="SISR_FM_AMZ.xlsm"
FM_path=folderpath+FM_SISR

SM_SISR="SISR_SM_AMZ.xlsm"
SM_path=folderpath+SM_SISR

PB_SISR="SISR_PB_AMZ.xlsm"
PB_path=folderpath+PB_SISR

OT_SISR="SISR_OT_AMZ.xlsm"
OT_path=folderpath+OT_SISR

FN_SISR="SISR_FN_AMZ.xlsm"
FN_path=folderpath+FN_SISR

#IM_SISR="SISR_IM_Direct.xlsm"
#IM_path=folderpath+IM_SISR

# Load the workbook for all categories with all formulas

wb_SB=load_workbook(SB_path)
sheet_SB=wb_SB.active

wb_BX=load_workbook(BX_path)
sheet_BX=wb_BX.active

wb_FR=load_workbook(FR_path)
sheet_FR=wb_FR.active

wb_FM=load_workbook(FM_path)
sheet_FM=wb_FM.active

wb_SM=load_workbook(SM_path)
sheet_SM=wb_SM.active

wb_PB=load_workbook(PB_path)
sheet_PB=wb_PB.active

wb_OT=load_workbook(OT_path)
sheet_OT=wb_OT.active

wb_FN=load_workbook(FN_path)
sheet_FN=wb_FN.active

## Load the workbook with data values only
wb_SB_data=load_workbook(SB_path,data_only=True)
sheet_SB_data=wb_SB_data.active

wb_BX_data=load_workbook(BX_path,data_only=True)
sheet_BX_data=wb_BX_data.active

wb_FR_data=load_workbook(FR_path,data_only=True)
sheet_FR_data=wb_FR_data.active

wb_FM_data=load_workbook(FM_path,data_only=True)
sheet_FM_data=wb_FM_data.active

wb_SM_data=load_workbook(SM_path,data_only=True)
sheet_SM_data=wb_SM_data.active

wb_PB_data=load_workbook(PB_path,data_only=True)
sheet_PB_data=wb_PB_data.active

wb_OT_data=load_workbook(OT_path,data_only=True)
sheet_OT_data=wb_OT_data.active

wb_FN_data=load_workbook(FN_path,data_only=True)
sheet_FN_data=wb_FN_data.active

#wb_IM=load_workbook(IM_path)
#sheet_IM=wb_IM.active

# modification & save the file 

def pasteInv(worksheet,worksheet_data,startcol):
    for col in range(startcol+1,startcol+25):
        ### West
        for row in range(22,worksheet.max_row,52):
            worksheet[get_column_letter(col)+str(row)]=worksheet_data[get_column_letter(col)+str(row)].value
        ### East
        for row in range(31, worksheet.max_row,52):
            worksheet[get_column_letter(col)+str(row)]=worksheet_data[get_column_letter(col)+str(row)].value

def deleteAMZ(worksheet,startcol):
    for col in range(startcol+1,startcol+25):
        ### for Amazon Direct
        for row in range(37,worksheet.max_row,52):
            worksheet[get_column_letter(col)+str(row)]=0
        ### for Amazon DDS
        for row in range(35, worksheet.max_row,52):
            worksheet[get_column_letter(col)+str(row)]=0

pasteInv(sheet_SB,sheet_SB_data,startCol)
pasteInv(sheet_BX,sheet_BX_data,startCol)
pasteInv(sheet_FR,sheet_FR_data,startCol)
pasteInv(sheet_FM,sheet_FM_data,startCol)
pasteInv(sheet_SM,sheet_SM_data,startCol)
pasteInv(sheet_PB,sheet_PB_data,startCol)
pasteInv(sheet_OT,sheet_OT_data,startCol)
pasteInv(sheet_FN,sheet_FN_data,startCol)

  
deleteAMZ(sheet_SB,startCol)          
deleteAMZ(sheet_BX,startCol)
deleteAMZ(sheet_FR,startCol)
deleteAMZ(sheet_FM,startCol)
deleteAMZ(sheet_SM,startCol)
deleteAMZ(sheet_PB,startCol)
deleteAMZ(sheet_OT,startCol)
deleteAMZ(sheet_FN,startCol)

#deleteDirect(sheet_IM,startCol


        
#for col in range(startCol,startCol+20):
 #   for row in range(37,sheet_IM.max_row,52):
 #       sheet_IM[get_column_letter(col)+str(row)]=0

os.chdir("C:/Users/Elvis Ma/Desktop/Weekly Work/space_planning")

wb_SB.save("SISR_SB_DDS.xlsx")
wb_BX.save("SISR_BX_DDS.xlsx")
wb_FR.save("SISR_FR_DDS.xlsx")
wb_FM.save("SISR_FM_DDS.xlsx")
wb_SM.save("SISR_SM_DDS.xlsx")
wb_PB.save("SISR_PB_DDS.xlsx")
wb_OT.save("SISR_OT_DDS.xlsx")
wb_FN.save("SISR_FN_DDS.xlsx")
#wb_IM.save("SISR_IM_DDS.xlsx")

