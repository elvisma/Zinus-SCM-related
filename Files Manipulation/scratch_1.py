import xlrd
import xlwt
import openpyxl as xl
import pyexcel as p

def convert_xls_xlsx(old_file, new_file):
    p.save_book_as(file_name=old_file,
                   dest_file_name= new_file)


def abc_category():
    wb = xl.load_workbook("WOS_Report.xlsx")
    sheet2 = wb["Data_AMOUNT"]

    #Copy sheet Report_Wayfair into workbook
    wb1 = xl.load_workbook("WOS_Report_CG.xlsx")
    sheet4 = wb1.worksheets[2]
    sheet5 = wb.create_sheet(title="Report_Wayfair")
    for row in sheet4:
        for cell in row:
            sheet5[cell.coordinate].value = cell.value

    #Filling empty cell with 0 on Data_Amount sheet to avoid calculation errors
    for row in range(1, sheet2.max_row+1):
        for col in range(1, sheet2.max_column+1):
            if sheet2.cell(row= row, column=col).value is None:
                sheet2.cell(row = row, column =col).value =0
                print ("filling empty cell with 0 at cell(row = %d, col = %d)" %(row, col))

    wb.create_sheet(title="Report_ZINUS")
    sheet3 = wb["Report_ZINUS"]

    new_list = abc+workbooks
    new_list = new_list + [m +'_'+ n for m in workbooks for n in abc]+['Total_ZINUS']
    print (new_list)

    #Read ETA from Sheet2
    for n in range(1,sheet2.max_column+1):
        sheet3.cell(row=1,column=n).value=sheet2.cell(row=1,column=n).value

    # adding frame
    for n in range(0, len(new_list)):
        row = n+1
        sheet3.cell(row=row * 4, column = 1).value = 'Total Import Plan'
        sheet3.cell(row=row * 4 + 1, column=1).value = 'Total Forecast Qty'
        sheet3.cell(row=row * 4 + 2, column=1).value = 'Total Ending Inventory'
        sheet3.cell(row=row * 4 + 3, column=1).value = 'WOS'
        sheet3.cell(row=row * 4 , column=2).value = new_list[n]
        sheet3.cell(row=row * 4 + 1, column=2).value = new_list[n]
        sheet3.cell(row=row * 4 + 2, column=2).value = new_list[n]
        sheet3.cell(row=row * 4 + 3, column=2).value = new_list[n]

    for colNum in range (15,sheet2.max_column+1):   # col 15 start with data

        # form dictionaries with a list starting with 0
        forecast = dict.fromkeys(new_list,0)
        ending = dict.fromkeys(new_list,0)
        importplan = dict.fromkeys(new_list, 0)

        for rowNum in range(2, sheet2.max_row+1):
            if sheet2.cell(row = rowNum, column=14).value == "Total Actual Qty" :
                abc_key = sheet2.cell(row = rowNum, column=1).value
                cate_key = sheet2.cell(row=rowNum, column=2).value
                if cate_key in new_list and abc_key in new_list:
                    forecast[abc_key] += float(sheet2.cell(row= rowNum, column = colNum).value)
                    forecast[cate_key] += float(sheet2.cell(row= rowNum, column = colNum).value)
                    forecast[cate_key+'_'+abc_key] += float(sheet2.cell(row=rowNum, column=colNum).value)
                    forecast['Total_ZINUS']+= float(sheet2.cell(row= rowNum, column = colNum).value)

            elif (sheet2.cell(row = rowNum, column=14).value == "Total Ending Inventory"):
                abc_key = sheet2.cell(row = rowNum, column=1).value
                cate_key = sheet2.cell(row=rowNum, column=2).value
                if cate_key in new_list and abc_key in new_list:
                    ending[abc_key] += float(sheet2.cell(row= rowNum, column = colNum).value)
                    ending[cate_key] += float(sheet2.cell(row= rowNum, column = colNum).value)
                    ending[cate_key + '_' + abc_key] += float(sheet2.cell(row=rowNum, column=colNum).value)
                    ending['Total_ZINUS']+= float(sheet2.cell(row= rowNum, column = colNum).value)

            elif (sheet2.cell(row = rowNum, column=14).value == "Total Import Plan"):
                abc_key = sheet2.cell(row = rowNum, column=1).value
                cate_key = sheet2.cell(row=rowNum, column=2).value
                if cate_key in new_list and abc_key in new_list:
                    importplan[abc_key] += float(sheet2.cell(row= rowNum, column = colNum).value)
                    importplan[cate_key] += float(sheet2.cell(row= rowNum, column = colNum).value)
                    importplan[cate_key + '_' + abc_key] += float(sheet2.cell(row=rowNum, column=colNum).value)
                    importplan['Total_ZINUS']+= float(sheet2.cell(row= rowNum, column = colNum).value)


        for n in range (len(new_list)):
            row = n+ 1
            sheet3.cell(row=row * 4, column=colNum).value =importplan[new_list[n]]
            sheet3.cell(row=row * 4 + 1, column=colNum).value = forecast[new_list[n]]
            sheet3.cell(row=row * 4 + 2, column=colNum).value = ending[new_list[n]]

        #Combine Zinus with Wayfair
        for cl in range (1, sheet5.max_row+1):
            if sheet3.cell(row=1, column=colNum).value == sheet5.cell(row=1,column=cl).value:
                sheet3.cell(row=row + 1, column=colNum).value = sheet5.cell(row=120, column=cl).value
                print(sheet3.cell(row=row + 1, column=colNum).value)

    for col in range (15, sheet3.max_column+1-7): # last 7 col won't be calculated
        for row in range (3, sheet3.max_row+1):
            if sheet3.cell(row = row, column = 1).value == "WOS":
                #print row, col
                try:
                    sheet3.cell(row = row, column = col).value = int(float(sheet3.cell(row = row-1, column = col).value)/((float (sheet3.cell(row = row-2, column = col).value)+float(sheet3.cell(row = row-2, column = col+1).value)+float(sheet3.cell(row = row-2, column = col+2).value)+ float(sheet3.cell(row = row-2, column = col+3).value)+float(sheet3.cell(row = row-2, column = col+4).value)+float(sheet3.cell(row = row-2, column = col+5).value)+float(sheet3.cell(row = row-2, column = col+6).value)+float(sheet3.cell(row = row-2, column = col+7).value))/8))
                except ZeroDivisionError:
                    print ("you can't devide by 0, filling WOS cell (row = %d, col =%d) failed!" %(row, col))
                    sheet3.cell(row=row, column=col).value =0
                    continue


    wb.save("WOS_Report.xlsx")
    print ("Program finished, new report generated!")

def zinus_sisr():
    #extract abc category info
    abc_dict= {}
    category_file = ['MATTRESS+AMOUNT', 'NEW+ITEM+MATTRESS+AMOUNT', 'NEW+ITEM+OTHERS+AMOUNT',
                     'NEW+ITEM+PLATFORM+BED+AMOUNT', 'FRAME+AMOUNT', 'PLATFORM+BED+AMOUNT', 'OTHERS+AMOUNT',
                     'NEW+ITEM+FRAME+AMOUNT', 'NEW+ITEM+BOX+SPRING+AMOUNT', 'BOX+SPRING+AMOUNT']
    for n in range(len(category_file)):
        file_name =category_file[n]+'.xlsx'  # Path to your file
        abc_file = xlrd.open_workbook(file_name, on_demand=True)
        abc_sheet = abc_file.sheet_by_index(0)
        for row in range(abc_sheet.nrows):
            read_row = abc_sheet.row_values(row)
            if read_row[0] not in abc_dict.keys():
                abc_dict[read_row[0]] = read_row[3]
            elif read_row[0] == 'ABC':
                print ("skip the header ABC in the first line of the file")
            else:
                print ("SKU %s got from file %s is already in the abc_dict, please check! " %(abc_dict[read_row[0]],category_file[n]))
    print ("abc category info extraction is done.")
    #######################################################################################

    # extract discontinue item list
    discont_dict = {}
    category_file = ['Discontinued_ItemList']
    for n in range(len(category_file)):
        file_name = category_file[n] + '.xlsx'  # Path to your file
        discont_file = xlrd.open_workbook(file_name, on_demand=True)
        discont_sheet = discont_file.sheet_by_index(0)
        for row in range(discont_sheet.nrows):
            read_row = discont_sheet.row_values(row)
            if read_row[0] not in discont_dict.keys():
                discont_dict[read_row[0]] = read_row[2]
            else:
                print("SKU %s got from file %s is already in the abc_dict, please check! " % (discont_dict[read_row[0]], category_file[n]))
    print("Discontinued item list info extraction is done.")

    # extract Inventory info
    keyword_row_num = {}  # dictionary
    keywords = ["ETA", "Total Actual Qty", "Total Ending Inventory","Total WOS","Total Import Plan","AMAZON-DDS","AMAZON-DDS","AMAZON-DIREC"]   #list
    #keywords = ["AMAZON-DDS", "COSTCO.com", "HOMEDEPT","OVERSTOCK","WMT.com","SAMCLU","TARGET","WAYFAIR","ZINUS.com"]   #list
    print ("Loading xlsm files...")
    workbook_date="12232018"

    workbook_wr = xlwt.Workbook()
    sheet_1 = workbook_wr.add_sheet("Data_QTY")
    sheet_2 = workbook_wr.add_sheet("Data_AMOUNT")
    row_cnt=0
    ETA_FLAG= 0

    #add new sheet for $ value per SKU
    workbook_sku_value = xlrd.open_workbook("FOB_Cost.xlsx", on_demand=True)
    sheet_sku_value = workbook_sku_value.sheet_by_index(0)

    #put {sku, value} pair into dictionary
    sku_dict = {}
    for row in range (sheet_sku_value.nrows):
        read_row = sheet_sku_value.row_values(row)
        sku_dict[read_row[0]] = read_row[1]
    #print sku_dict.keys()

    for n in range(len(workbooks)):
        workbook_name="SISR_"+workbooks[n]+"_"+workbook_date+".xlsm"
        workbook_0 = xlrd.open_workbook(workbook_name, on_demand=True)
        sheet = workbook_0.sheet_by_index(0)
        key_cnt = 0
        for key in keywords:
            row_num = 0
            key_cnt = key_cnt + 1
            if key == "ETA" and ETA_FLAG ==1:      # use flag to escape many ETA print
                continue
            #print ('Writing rows have keyword in %s: %s ..' %(workbook_name, key))
            for row in range(sheet.nrows):
                if key in sheet.row_values(row):
                    row_num = row_num + 1
                    col_cnt=0
                    #print (sheet.row_values(row))
                    read_row = sheet.row_values(row)
                    if workbooks[n]=='FN':
                        read_row.insert(0,'OT')
                    else:
                        read_row.insert(0, workbooks[n]) # add category name into front of each row
                    if read_row[7] in discont_dict.keys():
                        read_row.insert(0, discont_dict[read_row[7]])
                    elif read_row[7] in abc_dict.keys():
                        read_row.insert(0, abc_dict[read_row[7]])
                    else:
                        read_row.insert(0, 'New')
                    for col_cnt in range(len(read_row)):
                        if col_cnt ==0 and row_cnt == 0:
                            sheet_1.write(row_cnt, col_cnt, 'ABC')
                        elif col_cnt ==1 and row_cnt == 0:
                            sheet_1.write(row_cnt, col_cnt, '')
                        else:
                            sheet_1.write(row_cnt, col_cnt, read_row[col_cnt])
                        if col_cnt >= 14 and row_cnt > 0:  #14 is the start line with data
                            try :
                                int(read_row[col_cnt])
                            except ValueError:
                                print ('%s cannot convert to int' %read_row[col_cnt])
                                continue
                            # read_row[8] is the SKU number
                            if read_row[8] in sku_dict.keys():
                                sheet_2.write(row_cnt, col_cnt, read_row[col_cnt]*(sku_dict[read_row[8]]))
                                #print read_row[col_cnt]*(sku_dict[read_row[7]])
                            else:
                                print ('%s is not in the SKU dict!!! Please check!!!' %read_row[8])
                        else:
                            if col_cnt == 0 and row_cnt == 0:
                                sheet_2.write(row_cnt, col_cnt, 'ABC')
                            elif col_cnt == 1 and row_cnt == 0:
                                sheet_2.write(row_cnt, col_cnt, '')
                            else:
                                sheet_2.write(row_cnt, col_cnt, read_row[col_cnt])

                        col_cnt=col_cnt+1
                    row_cnt = row_cnt + 1
                    if key == "ETA":
                        ETA_FLAG = 1

                keyword_row_num[key]  = row_num



        print ("-------------------------------------------------------------------------")
        print ("Summary for %s.." %workbook_name)
        print ("There are %d keywords have been proccessed:"  %key_cnt)
        for x in keyword_row_num:
            print ("\'%s\' has %d rows" %(x, keyword_row_num[x]))
        print ("Writing Finished!")
        print ("-------------------------------------------------------------------------")


    workbook_wr.save("WOS_Report.xls")


    print ("New report generation finished!")





if __name__ == "__main__":
    #workbooks = ["FM"]  # use this to quickly try out code
    workbooks = ["FM","SM","BX","FR","PB","FN","OT" ]
    abc = ["A", "B", "C","NA", "NB", "NC","Discontinued","New"]
    #abc =  ["A","B", "C"]
    zinus_sisr()
    convert_xls_xlsx("WOS_Report.xls", "WOS_Report.xlsx")
    abc_category()